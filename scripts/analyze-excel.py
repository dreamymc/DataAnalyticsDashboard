#!/usr/bin/env python3
"""
analyze-excel.py — Quick analysis of the T8 dashboard Excel data.
Run: python3 scripts/analyze-excel.py
"""
import sys
import openpyxl
from collections import Counter

EXCEL_PATH = 'info/T8_Master_Dataset_Populated.xlsx'

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    col = {h: i for i, h in enumerate(headers) if h}

    print(f"=== T8 Master Dataset Analysis ===")
    print(f"Rows: {len(rows)}, Columns: {len(headers)}")
    print(f"Headers: {headers}")

    # Key computed metrics
    plan_rows = [r for r in rows if r[col.get('263 List / PLAN (In-Year)', -1)] == 'Yes']
    actual_rows = [r for r in rows if r[col.get('Actual Month (TRFS)', -1)] and r[col.get('Actual Month (TRFS)', -1)] != 'N/A']
    print(f"\n--- Scorecard Values ---")
    print(f"Pipeline (total): {len(rows)}")
    print(f"Plan (263 List=Yes): {len(plan_rows)}")
    print(f"Actual TRFS (has actual month): {len(actual_rows)}")
    print(f"%TRFS: {len(actual_rows)/len(plan_rows)*100:.1f}%")

    # High Level Status (Stage Funnel)
    hls = Counter(r[col['High Level Status']] for r in rows if r[col['High Level Status']])
    print(f"\n--- High Level Status (Stage Funnel) ---")
    for k, v in sorted(hls.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")

    # CW Status (RFI Rally)
    cws = Counter(r[col['CW Status']] for r in rows if r[col['CW Status']])
    print(f"\n--- CW Status (RFI Rally) ---")
    for k, v in sorted(cws.items(), key=lambda x: x[0]):
        print(f"  {k}: {v}")

    # Target Month (Build Plan)
    tm = Counter(r[col['Target Month (TRFS Plan)']] for r in plan_rows if r[col['Target Month (TRFS Plan)']] and r[col['Target Month (TRFS Plan)']] != 'N/A')
    am = Counter(r[col['Actual Month (TRFS)']] for r in actual_rows)
    print(f"\n--- Build Plan (monthly) ---")
    for month in ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC']:
        print(f"  {month}: plan={tm.get(month,0)}, actual={am.get(month,0)}")

if __name__ == '__main__':
    main()
